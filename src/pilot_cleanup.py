"""Pilot Cleanup — fix 6 issues before manual review.

1. Remove internal node IDs from all text fields
2. Rewrite negative_space questions into practitioner scenarios
3. Force cross_reference questions to one primary reference
4. Rename rag_likely_answer -> author_expected_failure
5. Add review_decision and review_notes fields
6. Export review XLSX for manual review
"""

from __future__ import annotations
import json, re, os, sys
from pathlib import Path
from copy import deepcopy

PILOT_PATH = Path(__file__).resolve().parent.parent / "data" / "pilot"
INPUT_FILE = PILOT_PATH / "treebench_pilot_100.json"
OUTPUT_JSON = PILOT_PATH / "treebench_pilot_100_cleaned.json"
OUTPUT_XLSX = PILOT_PATH / "treebench_pilot_review.xlsx"

# Pattern to catch any internal node IDs that might leak
NODE_ID_RE = re.compile(r"ECFR_TITLE\d+_XML__\S+")
# Pattern for raw set notation in negative_space signal text
SET_SIGNAL_RE = re.compile(r"""Covers\s*\{([^}]+)\},\s*missing\s*\{([^}]+)\}""")


# ──────────────────────────────────────────────────────────────────────
# Fix 1: Scrub internal node IDs from all text fields
# ──────────────────────────────────────────────────────────────────────
TEXT_FIELDS = ["question", "gold_answer", "rag_likely_answer", "why_similarity_fails"]


def fix_node_ids(q: dict) -> dict:
    """Remove internal node IDs like ECFR_TITLE26_XML__title_1__chapter_I..."""
    for field in TEXT_FIELDS:
        if field in q:
            q[field] = NODE_ID_RE.sub("[node]", q[field])
    # Also clean required_node_ids display (keep the list but don't expose in text)
    return q


# ──────────────────────────────────────────────────────────────────────
# Fix 2: Rewrite negative_space questions as practitioner scenarios
# ──────────────────────────────────────────────────────────────────────
NEGATIVE_SPACE_SCENARIOS = {
    # (covered_set_key, missing_term) -> (scenario_question, scenario_answer)
    ("citizen",): {
        "scenario": (
            "determines whether a specific provision exists for citizens. "
            "The section addresses residents and nonresidents but the practitioner "
            "needs guidance specifically applicable to U.S. citizens. Is there a "
            "dedicated provision for citizens within this regulatory subtree?"
        ),
        "answer_detail": (
            "covers residents and nonresidents but contains no dedicated provision "
            "for citizens. This gap means the practitioner must look elsewhere in the "
            "regulatory framework for citizen-specific rules, or determine whether "
            "the resident/nonresident provisions apply by analogy."
        ),
    },
    ("self-employed",): {
        "scenario": (
            "needs to determine the applicable rules for a self-employed individual. "
            "The section addresses employers and employees, but the practitioner's "
            "client is neither — they are self-employed. Is there a provision "
            "covering self-employed persons within this regulatory subtree?"
        ),
        "answer_detail": (
            "covers employers and employees but contains no dedicated provision "
            "for self-employed individuals. This regulatory gap means the self-employed "
            "person falls outside the explicit scope of this section and must seek "
            "guidance under a different part of the regulatory framework."
        ),
    },
    ("nonresident",): {
        "scenario": (
            "is reviewing the section for guidance on nonresidents. "
            "The provisions cover residents and citizens, but the client "
            "is a nonresident. Does this subtree contain a specific "
            "provision addressing nonresidents?"
        ),
        "answer_detail": (
            "covers residents and citizens but contains no dedicated provision "
            "for nonresidents. The practitioner must look to a different regulatory "
            "section for nonresident-specific rules."
        ),
    },
    ("estate", "trust"): {
        "scenario": (
            "needs to determine whether estates and trusts are covered. "
            "The section addresses individuals, corporations, and partnerships, "
            "but the client is an estate. Are estates and trusts addressed "
            "within this regulatory subtree?"
        ),
        "answer_detail": (
            "covers individuals, corporations, and partnerships but contains no "
            "dedicated provision for estates or trusts. These entity types must "
            "seek guidance under a different section of the regulatory framework."
        ),
    },
}


def _parse_signal(signal_text: str) -> tuple[set[str], set[str]]:
    """Parse 'Covers {a, b}, missing {c}' into sets."""
    m = SET_SIGNAL_RE.search(signal_text)
    if not m:
        return set(), set()
    covered_raw = m.group(1).replace("'", "").replace('"', '')
    missing_raw = m.group(2).replace("'", "").replace('"', '')
    covered = {s.strip() for s in covered_raw.split(",")}
    missing = {s.strip() for s in missing_raw.split(",")}
    return covered, missing


def fix_negative_space(q: dict) -> dict:
    """Rewrite negative_space questions from raw set notation to practitioner scenarios."""
    if q.get("failure_type") != "negative_space":
        return q

    signal = q.get("gold_answer", "") + q.get("question", "")
    covered, missing = _parse_signal(signal)

    if not missing:
        return q

    # Find matching scenario template
    missing_key = tuple(sorted(missing))
    scenario_data = NEGATIVE_SPACE_SCENARIOS.get(missing_key)

    if not scenario_data:
        # Fallback: generate from the missing terms directly
        missing_list = ", ".join(sorted(missing))
        covered_list = ", ".join(sorted(covered))
        scenario_data = {
            "scenario": (
                f"needs to determine whether {missing_list} "
                f"{'is' if len(missing) == 1 else 'are'} addressed in this regulatory section. "
                f"The section covers {covered_list}, but the practitioner's specific "
                f"scenario involves {missing_list}. Is there a dedicated provision?"
            ),
            "answer_detail": (
                f"covers {covered_list} but contains no dedicated provision "
                f"for {missing_list}. This regulatory gap means the practitioner "
                f"must seek guidance under a different part of the framework."
            ),
        }

    # Extract source and heading info from existing question
    source_title = q.get("source_title", "the regulation")
    heading = q.get("gold_path", [""])[-1] if q.get("gold_path") else "this section"
    tree_path = " > ".join(q.get("gold_path", []))

    # Rewrite question
    q["question"] = (
        f"A practitioner reviewing {source_title} at {heading} "
        f"{scenario_data['scenario']}"
    )

    # Rewrite gold answer
    covered_list = ", ".join(sorted(covered))
    missing_list = ", ".join(sorted(missing))
    q["gold_answer"] = (
        f"No. The regulatory subtree at {tree_path} "
        f"{scenario_data['answer_detail']}"
    )

    # Rewrite rag_likely_answer
    q["rag_likely_answer"] = (
        f"A similarity-based retrieval system would return the nearest section "
        f"covering {covered_list}, which discusses related entity types. This "
        f"produces a confident but incorrect answer that assumes {missing_list} "
        f"{'is' if len(missing) == 1 else 'are'} covered under the same provision. "
        f"The system cannot detect that the subtree has no matching node for "
        f"{missing_list} — it always returns something."
    )

    # Clean up why_similarity_fails
    q["why_similarity_fails"] = (
        f"Negative space — the absence of a provision for {missing_list} — cannot "
        f"be detected by cosine similarity. The query about {missing_list} has high "
        f"semantic overlap with the existing provisions for {covered_list}, so the "
        f"retrieval system returns a related section with high confidence. Only an "
        f"exhaustive tree traversal can confirm that no matching child node exists."
    )

    return q


# ──────────────────────────────────────────────────────────────────────
# Fix 3: Force cross_reference to one primary reference
# ──────────────────────────────────────────────────────────────────────
def fix_cross_reference(q: dict) -> dict:
    """Reduce cross_reference questions to one primary reference."""
    if q.get("failure_type") != "cross_reference":
        return q

    # Find all section references in the question
    refs = re.findall(
        r"(?:Section|section|sec\.|§)\s*([\d]+(?:\.[\d\w\-]+)*(?:\([a-zA-Z0-9]+\))*)",
        q["question"]
    )

    if len(refs) <= 1:
        return q

    # Keep only the first reference
    primary_ref = refs[0]
    primary_full = f"Section {primary_ref}"

    # Get source/heading info
    source_title = q.get("source_title", "the regulation")
    heading = q.get("gold_path", [""])[-1] if q.get("gold_path") else "this section"
    tree_path = " > ".join(q.get("gold_path", []))

    # Rewrite question with single reference
    q["question"] = (
        f"A regulatory analyst reviewing {heading} under {source_title} "
        f"encounters a cross-reference to {primary_full}. What does the "
        f"cross-referenced provision provide, and how does it modify or "
        f"qualify the rule at {heading}?"
    )

    # Update gold answer to focus on single reference
    gold_answer = q.get("gold_answer", "")
    # Replace multi-ref mentions with single ref
    q["gold_answer"] = re.sub(
        r"(?:Section|section|sec\.|§)\s*[\d]+(?:\.[\d\w\-]+)*"
        r"(?:\s*,\s*(?:Section|section|sec\.|§)\s*[\d]+(?:\.[\d\w\-]+)*)+",
        primary_full,
        gold_answer,
        count=1,
    )

    # Update rag_likely_answer
    q["rag_likely_answer"] = (
        f"Similarity-based retrieval returns the section containing the "
        f"cross-reference to {primary_full}, but does NOT follow the link "
        f"to retrieve the referenced provision's content. The system treats "
        f"the reference text as content rather than as a traversal instruction, "
        f"so the analyst receives the rule that mentions {primary_full} but "
        f"not the substance of what {primary_full} actually provides."
    )

    return q


# ──────────────────────────────────────────────────────────────────────
# Fix 4: Rename rag_likely_answer -> author_expected_failure
# Fix 5: Add review_decision and review_notes
# ──────────────────────────────────────────────────────────────────────
def fix_schema(q: dict) -> dict:
    """Rename field and add review fields."""
    # Fix 4: Rename
    if "rag_likely_answer" in q:
        q["author_expected_failure"] = q.pop("rag_likely_answer")

    # Fix 5: Add review fields
    if "review_decision" not in q:
        q["review_decision"] = ""
    if "review_notes" not in q:
        q["review_notes"] = ""

    return q


# ──────────────────────────────────────────────────────────────────────
# Fix 6: Export review XLSX
# ──────────────────────────────────────────────────────────────────────
def export_review_xlsx(questions: list[dict], xlsx_path: Path) -> None:
    """Export a review-friendly XLSX with key columns for manual review."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    # Column order optimized for review workflow
    review_cols = [
        ("question_id", 18),
        ("domain", 12),
        ("failure_type", 22),
        ("structural_confounder_type", 18),
        ("difficulty", 10),
        ("tree_depth", 10),
        ("question", 60),
        ("gold_answer", 60),
        ("author_expected_failure", 60),
        ("why_similarity_fails", 50),
        ("gold_path", 40),
        ("answer_type", 15),
        ("source_title", 25),
        ("review_decision", 15),
        ("review_notes", 40),
        ("review_status", 18),
    ]

    # Headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for col_idx, (col_name, width) in enumerate(review_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Data rows
    wrap_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for row_idx, q in enumerate(questions, 2):
        for col_idx, (col_name, _) in enumerate(review_cols, 1):
            value = q.get(col_name, "")
            if isinstance(value, list):
                value = " > ".join(str(v) for v in value) if col_name == "gold_path" else json.dumps(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=str(value)[:32000])  # Excel cell limit
            cell.alignment = wrap_align
            cell.border = thin_border

            # Highlight review columns
            if col_name in ("review_decision", "review_notes"):
                cell.fill = review_fill

    # Freeze panes (header + question_id visible while scrolling)
    ws.freeze_panes = "G2"

    # Matrix sheet — coverage summary
    ws2 = wb.create_sheet("Coverage")
    ws2.append(["Domain", "Failure Type", "Count", "Avg Depth", "Difficulties"])

    from collections import Counter, defaultdict
    cell_data = defaultdict(list)
    for q in questions:
        cell_data[(q["domain"], q["failure_type"])].append(q)

    for (domain, ftype), cell_qs in sorted(cell_data.items()):
        avg_depth = sum(q["tree_depth"] for q in cell_qs) / len(cell_qs)
        diffs = Counter(q["difficulty"] for q in cell_qs)
        diff_str = ", ".join(f"{d}:{c}" for d, c in sorted(diffs.items()))
        ws2.append([domain, ftype, len(cell_qs), round(avg_depth, 1), diff_str])

    wb.save(str(xlsx_path))


# ──────────────────────────────────────────────────────────────────────
# Main: Apply all fixes
# ──────────────────────────────────────────────────────────────────────
def run_cleanup():
    print("TreeBench Pilot Cleanup")
    print("=" * 60)

    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        questions = json.load(f)

    print(f"Loaded {len(questions)} questions from {INPUT_FILE.name}")

    # Apply fixes in order
    fixed = []
    fix_counts = {"node_ids": 0, "negative_space": 0, "cross_ref": 0, "schema": 0}

    for q in questions:
        q_orig = deepcopy(q)

        # Fix 1: Node IDs
        q = fix_node_ids(q)
        if q != q_orig:
            fix_counts["node_ids"] += 1

        # Fix 2: Negative space
        q_before = deepcopy(q)
        q = fix_negative_space(q)
        if q["question"] != q_before.get("question"):
            fix_counts["negative_space"] += 1

        # Fix 3: Cross reference
        q_before = deepcopy(q)
        q = fix_cross_reference(q)
        if q["question"] != q_before.get("question"):
            fix_counts["cross_ref"] += 1

        # Fix 4 + 5: Schema
        q = fix_schema(q)
        fix_counts["schema"] += 1

        fixed.append(q)

    # Save cleaned JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(fixed, f, indent=2, ensure_ascii=False)
    print(f"\nCleaned JSON saved: {OUTPUT_JSON.name}")

    # Fix 6: Export XLSX
    export_review_xlsx(fixed, OUTPUT_XLSX)
    print(f"Review XLSX saved: {OUTPUT_XLSX.name}")

    # Summary
    print(f"\nFixes applied:")
    print(f"  Node IDs scrubbed:      {fix_counts['node_ids']}")
    print(f"  Negative space rewrite: {fix_counts['negative_space']}")
    print(f"  Cross-ref simplified:   {fix_counts['cross_ref']}")
    print(f"  Schema updated:         {fix_counts['schema']}")

    # Verify field presence
    sample = fixed[0]
    print(f"\nSchema verification:")
    print(f"  Has author_expected_failure: {'author_expected_failure' in sample}")
    print(f"  Has rag_likely_answer:       {'rag_likely_answer' in sample}")
    print(f"  Has review_decision:         {'review_decision' in sample}")
    print(f"  Has review_notes:            {'review_notes' in sample}")

    # Show a cleaned negative_space example
    print(f"\n--- Sample cleaned NEGATIVE_SPACE ---")
    for q in fixed:
        if q["failure_type"] == "negative_space":
            print(f"  Q: {q['question'][:200]}")
            print(f"  A: {q['gold_answer'][:200]}")
            print(f"  FAIL: {q['author_expected_failure'][:200]}")
            break

    # Show a cleaned cross_reference example
    print(f"\n--- Sample cleaned CROSS_REFERENCE ---")
    for q in fixed:
        if q["failure_type"] == "cross_reference":
            print(f"  Q: {q['question'][:200]}")
            break

    return fixed


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    run_cleanup()
