"""Balanced Sampler — select 1,000 questions for TreeBench final dataset.

Target distribution: 5 domains x 10 failure types = 50 cells
  1000 / 50 = 20 questions per cell

Difficulty distribution per cell: ~7 easy, ~8 medium, ~5 hard

Prioritization within each cell:
  1. Higher confidence candidates first
  2. Deeper tree paths preferred (more structurally interesting)
  3. Questions with cross-references get bonus (more complex reasoning)
"""

from __future__ import annotations
import json, random, os
from dataclasses import asdict
from collections import defaultdict
from typing import Optional
from question_generator import TreeBenchQuestion


TARGET_TOTAL = 1000
DOMAINS = ["tax", "finance", "medical", "legal", "compliance"]
FAILURE_TYPES = [
    "override_chain", "scope_disambiguation", "cross_reference",
    "conditional_cascade", "temporal_layering", "sibling_conflict",
    "definitional_dependency", "aggregation", "negative_space",
    "depth_gated_specificity",
]
DIFFICULTY_DIST = {"easy": 0.35, "medium": 0.40, "hard": 0.25}


def _quality_score(q: TreeBenchQuestion) -> float:
    """Score a question for sampling priority. Higher = better."""
    score = 0.0
    # Deeper tree paths are more interesting
    score += min(q.tree_depth / 8.0, 1.0) * 30

    # Longer questions are usually more specific
    score += min(len(q.question) / 300, 1.0) * 20

    # Having a secondary failure type indicates complexity
    if q.failure_type_secondary:
        score += 15

    # More reasoning steps = more valuable
    score += min(q.reasoning_steps_required / 5, 1.0) * 20

    # Cross-references in gold evidence
    evidence = q.gold_evidence
    if "section" in evidence.lower() or "§" in evidence:
        score += 10

    # Penalize very short answers
    if len(q.correct_answer) < 50:
        score -= 20

    return score


def sample_balanced(all_questions: list[TreeBenchQuestion],
                    target: int = TARGET_TOTAL,
                    seed: int = 42) -> list[TreeBenchQuestion]:
    """Sample a balanced dataset from all generated questions."""
    rng = random.Random(seed)

    # Group by (domain, failure_type)
    cells: dict[tuple[str, str], list[TreeBenchQuestion]] = defaultdict(list)
    for q in all_questions:
        key = (q.domain, q.failure_type_primary)
        cells[key].append(q)

    # Sort each cell by quality score (descending)
    for key in cells:
        cells[key].sort(key=_quality_score, reverse=True)

    # Calculate per-cell target
    num_cells = len(DOMAINS) * len(FAILURE_TYPES)  # 50
    per_cell = target // num_cells  # 20

    sampled: list[TreeBenchQuestion] = []
    cell_stats: dict[str, int] = {}
    shortfall_cells: list[tuple[str, str]] = []

    for domain in DOMAINS:
        for ftype in FAILURE_TYPES:
            key = (domain, ftype)
            available = cells.get(key, [])
            cell_label = f"{domain}/{ftype}"

            if len(available) == 0:
                shortfall_cells.append(key)
                cell_stats[cell_label] = 0
                continue

            # Try to hit difficulty distribution within cell
            by_diff = defaultdict(list)
            for q in available:
                by_diff[q.difficulty].append(q)

            cell_sample: list[TreeBenchQuestion] = []
            for diff, ratio in DIFFICULTY_DIST.items():
                n_want = max(1, round(per_cell * ratio))
                pool = by_diff.get(diff, [])
                take = min(n_want, len(pool))
                cell_sample.extend(pool[:take])

            # If we have fewer than per_cell, fill from remaining
            used_ids = {q.id for q in cell_sample}
            remaining = [q for q in available if q.id not in used_ids]
            while len(cell_sample) < per_cell and remaining:
                cell_sample.append(remaining.pop(0))

            # If still short, note it
            if len(cell_sample) < per_cell:
                shortfall_cells.append(key)

            sampled.extend(cell_sample[:per_cell])
            cell_stats[cell_label] = len(cell_sample[:per_cell])

    # Fill remaining slots from best available across all cells
    target_remaining = target - len(sampled)
    if target_remaining > 0:
        used_ids = {q.id for q in sampled}
        remaining_pool = [q for q in all_questions if q.id not in used_ids]
        remaining_pool.sort(key=_quality_score, reverse=True)
        sampled.extend(remaining_pool[:target_remaining])

    # Re-number IDs sequentially
    for i, q in enumerate(sampled):
        domain_abbrev = q.domain.upper()[:3]
        ftype_abbrev = q.failure_type_primary.upper().replace("_", "-")
        q.id = f"TB-{domain_abbrev}-{ftype_abbrev}-{i+1:04d}"

    return sampled


def export_dataset(questions: list[TreeBenchQuestion], output_dir: str) -> dict:
    """Export the final TreeBench dataset as JSON and XLSX."""
    os.makedirs(output_dir, exist_ok=True)

    # JSON export
    json_path = os.path.join(output_dir, "treebench_v1.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump([asdict(q) for q in questions], f, indent=2, ensure_ascii=False)

    # XLSX export
    try:
        import openpyxl
        wb = openpyxl.Workbook()

        # Dataset sheet
        ws = wb.active
        ws.title = "Dataset"
        headers = list(asdict(questions[0]).keys())
        ws.append(headers)
        for q in questions:
            ws.append(list(asdict(q).values()))

        # Matrix sheet — domain x failure_type counts
        ws2 = wb.create_sheet("Matrix")
        ws2.append(["domain"] + FAILURE_TYPES + ["total", "easy", "medium", "hard"])
        from collections import Counter
        for domain in DOMAINS:
            row = [domain]
            domain_qs = [q for q in questions if q.domain == domain]
            for ftype in FAILURE_TYPES:
                count = sum(1 for q in domain_qs if q.failure_type_primary == ftype)
                row.append(count)
            row.append(len(domain_qs))
            diff_counts = Counter(q.difficulty for q in domain_qs)
            row.extend([diff_counts.get("easy", 0), diff_counts.get("medium", 0), diff_counts.get("hard", 0)])
            ws2.append(row)
        # Total row
        total_row = ["TOTAL"]
        for ftype in FAILURE_TYPES:
            total_row.append(sum(1 for q in questions if q.failure_type_primary == ftype))
        total_row.append(len(questions))
        diff_all = Counter(q.difficulty for q in questions)
        total_row.extend([diff_all.get("easy", 0), diff_all.get("medium", 0), diff_all.get("hard", 0)])
        ws2.append(total_row)

        # Taxonomy sheet
        ws3 = wb.create_sheet("Taxonomy")
        ws3.append(["failure_type", "label", "definition", "rag_failure_pattern"])
        taxonomy_defs = {
            "override_chain": ("Override Chain", "A broad parent rule appears to answer the question, but a narrower child exception reverses or qualifies it.", "Retrieves the parent rule and misses the child exception."),
            "scope_disambiguation": ("Scope Disambiguation", "The same term is defined differently in multiple subtrees; the correct definition depends on which subtree the query falls in.", "Retrieves the most common definition, not the one scoped to the relevant subtree."),
            "cross_reference": ("Cross-Reference Traversal", "The answer requires following an explicit cross-reference to a different section/subtree.", "Retrieves the source section but does not follow the cross-reference link."),
            "conditional_cascade": ("Conditional Cascade", "The answer depends on a chain of 3+ conditions spread across multiple tree levels.", "Retrieves the leaf condition but misses ancestor conditions that gate it."),
            "temporal_layering": ("Temporal Layering", "The correct answer depends on a date-specific version or effective-date qualifier.", "Retrieves current text without surfacing the temporal qualifier."),
            "sibling_conflict": ("Sibling Conflict", "Two sibling sections at the same tree level provide seemingly conflicting rules.", "Retrieves the most similar sibling, which may state the wrong rule."),
            "definitional_dependency": ("Definitional Dependency", "The answer depends on a term whose definition is in a different subtree.", "Retrieves the rule using the term but not the definition that controls its meaning."),
            "aggregation": ("Aggregation Across Branches", "The answer requires combining values or rules from multiple separate subtrees.", "Retrieves one branch of the aggregation, missing the others."),
            "negative_space": ("Negative Space", "The correct answer is that the topic is NOT covered in the relevant subtree.", "Retrieves the nearest section by similarity, producing a false-positive answer."),
            "depth_gated_specificity": ("Depth-Gated Specificity", "A specific value/rate/threshold appears only at leaf level, not in any parent summary.", "Retrieves the parent summary, missing the specific value at depth."),
        }
        for ftype in FAILURE_TYPES:
            label, defn, pattern = taxonomy_defs[ftype]
            ws3.append([ftype, label, defn, pattern])

        xlsx_path = os.path.join(output_dir, "treebench_v1.xlsx")
        wb.save(xlsx_path)
        print(f"  XLSX saved: {xlsx_path}")
    except ImportError:
        print("  openpyxl not available, XLSX export skipped")
        xlsx_path = ""

    # Stats
    from collections import Counter
    stats = {
        "total_questions": len(questions),
        "by_domain": dict(Counter(q.domain for q in questions)),
        "by_failure_type": dict(Counter(q.failure_type_primary for q in questions)),
        "by_difficulty": dict(Counter(q.difficulty for q in questions)),
        "avg_tree_depth": round(sum(q.tree_depth for q in questions) / len(questions), 2),
        "avg_reasoning_steps": round(sum(q.reasoning_steps_required for q in questions) / len(questions), 2),
    }

    stats_path = os.path.join(output_dir, "treebench_v1_stats.json")
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2)

    print(f"  JSON saved: {json_path}")
    print(f"  Stats saved: {stats_path}")
    return stats


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    gen_dir = "../data/generated"
    gen_file = f"{gen_dir}/treebench_all_generated.json"

    if not os.path.exists(gen_file):
        print(f"Run question_generator.py first to create {gen_file}")
        sys.exit(1)

    # Load all generated questions
    with open(gen_file, "r", encoding="utf-8") as f:
        all_q_dicts = json.load(f)
    all_questions = [TreeBenchQuestion(**d) for d in all_q_dicts]
    print(f"Loaded {len(all_questions)} generated questions")

    # Sample balanced 1000
    sampled = sample_balanced(all_questions, target=1000)
    print(f"Sampled {len(sampled)} balanced questions")

    # Export
    output_dir = "../data/final"
    stats = export_dataset(sampled, output_dir)

    print(f"\n=== FINAL TREEBENCH v1 ===")
    print(f"Total: {stats['total_questions']}")
    print(f"\nBy domain:")
    for d, c in sorted(stats["by_domain"].items()):
        print(f"  {d:15s}: {c}")
    print(f"\nBy failure type:")
    for t, c in sorted(stats["by_failure_type"].items()):
        print(f"  {t:30s}: {c}")
    print(f"\nBy difficulty:")
    for d, c in sorted(stats["by_difficulty"].items()):
        print(f"  {d:10s}: {c}")
    print(f"\nAvg tree depth: {stats['avg_tree_depth']}")
    print(f"Avg reasoning steps: {stats['avg_reasoning_steps']}")
